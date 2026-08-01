from functools import lru_cache
from io import BytesIO
from typing import Any
from urllib.parse import quote, urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from core.logging_config import get_logger
from sources.parser import extract_dosage_form, extract_pack_size, extract_strength


EMA_BASE_URL = "https://www.ema.europa.eu"
EMA_MEDICINES_JSON_URL = (
    f"{EMA_BASE_URL}/en/documents/report/"
    "medicines-output-medicines_json-report_en.json"
)
EMA_MEDICINES_XLSX_URL = (
    f"{EMA_BASE_URL}/en/documents/report/"
    "medicines-output-medicines-report_en.xlsx"
)
EU_COUNTRIES = [
    "Austria",
    "Belgium",
    "Bulgaria",
    "Croatia",
    "Cyprus",
    "Czech Republic",
    "Denmark",
    "Estonia",
    "Finland",
    "France",
    "Germany",
    "Greece",
    "Hungary",
    "Ireland",
    "Italy",
    "Latvia",
    "Lithuania",
    "Luxembourg",
    "Malta",
    "Netherlands",
    "Poland",
    "Portugal",
    "Romania",
    "Slovakia",
    "Slovenia",
    "Spain",
    "Sweden",
]
EEA_COUNTRIES = ["Norway", "Iceland", "Liechtenstein"]
EMA_COUNTRIES = EU_COUNTRIES + EEA_COUNTRIES
logger = get_logger(__name__)


def _record_value(record: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = record.get(key)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _ema_product_url(record: dict[str, Any]) -> str:
    url = _record_value(
        record,
        "medicine_url",
        "url",
        "product_url",
        "epar_url",
    )
    return urljoin(EMA_BASE_URL, url) if url else ""


def _ema_holder(record: dict[str, Any]) -> str:
    return _record_value(
        record,
        "marketing_authorisation_developer_applicant_holder",
        "marketing_authorisation_holder_company_name",
        "marketing_authorisation_holder",
        "applicant",
        "holder",
    )


def _ema_strength(record: dict[str, Any], product: str) -> str:
    text = _record_value(
        record,
        "strength",
        "strengths",
        "active_substance_strength",
        "pharmaceutical_strength",
    )
    return text or extract_strength(product)


def _ema_dosage_form(record: dict[str, Any], product: str) -> str:
    text = _record_value(
        record,
        "pharmaceutical_form",
        "pharmaceutical_forms",
        "pharmaceutical_form_human",
        "dosage_form",
    )
    return text or extract_dosage_form(product)


def _ema_result_from_record(
    record: dict[str, Any],
    substance: str,
    country: str,
    source_url: str,
) -> dict[str, str]:
    product = _record_value(record, "name_of_medicine", "medicine_name", "product_name", "name")
    product_url = _ema_product_url(record)
    registration_date = _record_value(
        record,
        "marketing_authorisation_date",
        "european_commission_decision_date",
        "authorisation_date",
    )
    return {
        "substance": _record_value(record, "active_substance") or substance,
        "product": product,
        "company": _ema_holder(record),
        "country": country,
        "region": "EU",
        "status": _record_value(record, "medicine_status", "status"),
        "strength": _ema_strength(record, product),
        "dosage_form": _ema_dosage_form(record, product),
        "pack_size": extract_pack_size(product),
        "registration_number": _record_value(record, "ema_product_number", "product_number"),
        "registration_date": registration_date,
        "atc_code": _record_value(record, "atc_code_human", "atc_code"),
        "therapeutic_category": _record_value(
            record,
            "pharmacotherapeutic_group_human",
            "therapeutic_area",
            "therapeutic_indication",
        ),
        "source": "EMA",
        "source_url": source_url,
        "product_url": product_url,
        "url": product_url,
        "last_checked": _record_value(record, "last_updated_date", "last_updated"),
    }


def _extract_ema_products(html, substance, source_url):
    soup = BeautifulSoup(html, "html.parser")
    products = []
    seen_urls = set()
    for link in soup.select('a[href*="/en/medicines/human/EPAR/"]'):
        href = urljoin(EMA_BASE_URL, link.get("href", ""))
        product = " ".join(link.get_text(" ", strip=True).split())
        if not product or href in seen_urls:
            continue
        if substance.lower() not in product.lower():
            continue
        seen_urls.add(href)
        products.append({"product": product, "url": href, "source_url": source_url})
    return products


def _expand_to_ema_countries(products, substance):
    results = []
    for product in products:
        for country in EMA_COUNTRIES:
            results.append(
                {
                    "substance": substance,
                    "product": product["product"],
                    "company": "",
                    "country": country,
                    "region": "EU",
                    "status": "Authorised",
                    "strength": extract_strength(product["product"]),
                    "dosage_form": extract_dosage_form(product["product"]),
                    "source": "EMA",
                    "source_url": product["source_url"],
                    "product_url": product["url"],
                    "url": product["url"],
                }
            )
    return results


def _fetch_ema_medicine_json() -> list[dict[str, Any]]:
    response = requests.get(
        EMA_MEDICINES_JSON_URL,
        timeout=60,
        headers={"User-Agent": "Mozilla/5.0"},
    )
    response.raise_for_status()
    payload = response.json()
    return payload.get("data", [])


@lru_cache(maxsize=1)
def _fetch_ema_medicine_xlsx():
    response = requests.get(
        EMA_MEDICINES_XLSX_URL,
        timeout=60,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": f"{EMA_BASE_URL}/en/medicines/download-medicine-data",
        },
    )
    response.raise_for_status()
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = None
    records = []
    for row in worksheet.iter_rows(values_only=True):
        values = ["" if value is None else str(value).strip() for value in row]
        if not any(values):
            continue
        if values[0] == "Category":
            header = values
            continue
        if not header:
            continue
        record = {}
        for index, name in enumerate(header):
            if name:
                key = (
                    name.lower()
                    .replace("\n", " ")
                    .replace(" / ", " ")
                    .replace("/", " ")
                    .replace("(", "")
                    .replace(")", "")
                    .replace(":", "")
                    .replace("-", "_")
                    .replace(" ", "_")
                )
                record[key] = values[index] if index < len(values) else ""
        records.append(record)
    return records


def _record_matches_substance(record: dict[str, Any], substance: str) -> bool:
    query = substance.strip().lower()
    fields = [
        record.get("active_substance", ""),
        record.get("international_non_proprietary_name_common_name", ""),
        record.get("international_non_proprietary_name_inn_common_name", ""),
        record.get("name_of_medicine", ""),
    ]
    return any(query in str(field).lower() for field in fields)


def _expand_records(
    records: list[dict[str, Any]],
    substance: str,
    countries: list[str],
    source_url: str,
) -> list[dict[str, str]]:
    results = []
    for record in records:
        if not _record_matches_substance(record, substance):
            continue
        for country in countries:
            results.append(_ema_result_from_record(record, substance, country, source_url))
    return results


def _expand_json_records(records: list[dict[str, Any]], substance: str) -> list[dict[str, str]]:
    return _expand_records(records, substance, EMA_COUNTRIES, EMA_MEDICINES_JSON_URL)


def _expand_xlsx_records(records: list[dict[str, Any]], substance: str) -> list[dict[str, str]]:
    return _expand_records(records, substance, EU_COUNTRIES, EMA_MEDICINES_XLSX_URL)


def run_ema_search(substance: str) -> list[dict[str, str]]:
    try:
        return _expand_xlsx_records(_fetch_ema_medicine_xlsx(), substance)
    except (requests.RequestException, ValueError) as exc:
        logger.warning("EMA XLSX request failed: %s", exc)

    try:
        return _expand_json_records(_fetch_ema_medicine_json(), substance)
    except (requests.RequestException, ValueError) as exc:
        logger.warning("EMA JSON request failed: %s", exc)

    url = f"{EMA_BASE_URL}/en/search?search_api_fulltext={quote(substance)}"
    try:
        response = requests.get(
            url,
            timeout=45,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/137.0.0.0 Safari/537.36"
                )
            },
        )
    except requests.RequestException as exc:
        logger.warning("EMA search request failed: %s", exc)
        return []

    if response.status_code >= 500:
        logger.warning("EMA search request failed with status %s", response.status_code)
        return []

    if "technical difficulties with our search function" in response.text.lower():
        logger.warning("EMA search page is unavailable")
        return []
    products = _extract_ema_products(response.text, substance, url)
    return _expand_to_ema_countries(products, substance)


def find_product_url(substance):

    results = run_ema_search(substance)

    if not results:
        return None

    return results[0]["url"]

