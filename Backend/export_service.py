from datetime import datetime
from pathlib import Path
from collections import Counter
import re
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import EXPORT_DIR
from core.logging_config import get_logger
from repository import region_for_country
from services.ai_enrichment import enrichment_metadata
from services.english_normalizer import english_row
from services.field_availability import field_value
from services.result_formatter import (
    assessment_report_url as formatted_assessment_report_url,
    clean_brand_name,
    company_display_value,
    manufacturer_country_value,
    manufacturer_name_value,
    pil_or_assessment_url,
    product_details_url,
    smpc_document_url,
)
from services.therapeutic_category import (
    short_therapeutic_category,
    therapeutic_category_from_atc,
    therapeutic_category_from_substance,
)
from sources.parser import (
    extract_dosage_form,
    extract_pack_size,
    extract_registration_number,
    extract_route,
    extract_strength,
)


EXPORT_COLUMNS = [
    "Country",
    "Region",
    "Product Name",
    "Company",
    "Brand Name",
    "Molecule",
    "Molecule (Active Ingredient(s))",
    "Strength",
    "Dosage Form",
    "Route",
    "Pack Size",
    "ATC Code",
    "Therapeutic Category",
    "MA Holder Name",
    "Manufacturer Name",
    "Manufacturer Country",
    "Registration Status",
    "Registration Number",
    "Registration Date",
    "Expiry Date",
    "Product Details",
    "PDF URL",
    "SMPC URL",
    "PIL URL",
    "PIL / Assessment Report",
    "Assessment Report URL",
    "Manufacturer Website",
    "Manufacturer Contact Us Phone Number",
    "Manufacturer Contact Us Email ID",
    "Box Artwork",
    "Foil Artwork",
    "Insert / PIL artwork",
    "SMPC",
    "Source",
    "Source URL",
    "Document Type",
    "Data Confidence",
    "Enrichment Status",
    "Missing Fields",
    "AI Next Action",
    "Last Checked",
]
MISSING_VALUE = "Not available"
logger = get_logger(__name__)


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    return cleaned.strip("_") or "export"


def first_value(*values: object) -> str:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def export_value(*values: object) -> str:
    return first_value(*values) or MISSING_VALUE


def default_therapeutic_category(substance: str, results: list[dict[str, Any]]) -> str:
    values = []
    for item in results:
        category = first_value(
            short_therapeutic_category(
                item.get("therapeutic_category"),
                item.get("searched_substance") or item.get("substance") or substance,
                item.get("atc_code"),
            ),
            therapeutic_category_from_atc(item.get("atc_code")),
        )
        if category:
            values.append(category)
    if values:
        return Counter(values).most_common(1)[0][0]
    return therapeutic_category_from_substance(substance)


def infer_document_type(item: dict[str, Any]) -> str:
    document_type = item.get("document_type", "")
    if document_type:
        return document_type
    if item.get("smpc_url"):
        return "SPC"
    if item.get("pil_url"):
        return "PIL"
    if item.get("assessment_report_url"):
        return "Assessment report"
    return ""


def build_export_rows(substance: str, results: list[dict[str, Any]]) -> list[dict[str, str]]:
    rows = []
    generated_at = datetime.now().isoformat(timespec="seconds")
    default_category = default_therapeutic_category(substance, results)
    for item in results:
        if (
            str(item.get("connector_mode") or "").strip() == "manual_registry"
            or "registry search handoff" in str(item.get("document_type") or "").lower()
        ):
            continue
        item = english_row(item)
        ai_metadata = enrichment_metadata(item)
        product = clean_brand_name(item.get("product", ""))
        country = item.get("country", "")
        company = item.get("company", "")
        product_url = product_details_url(item)
        source_url = item.get("source_url") or item.get("url", "")
        pdf_url = first_value(item.get("document_url"), item.get("pdf_url"), item.get("url"))
        smpc_url = smpc_document_url(item)
        pil_url = item.get("pil_url", "")
        pil_or_assessment_document_url = pil_or_assessment_url(item)
        assessment_document_url = formatted_assessment_report_url(item)
        rows.append(
            {
                "Country": export_value(country),
                "Region": export_value(item.get("region"), region_for_country(country)),
                "Product Name": export_value(product),
                "Company": export_value(company_display_value(item)),
                "Brand Name": export_value(product),
                "Molecule": export_value(
                    item.get("searched_substance"),
                    item.get("substance"),
                    substance,
                ),
                "Molecule (Active Ingredient(s))": export_value(
                    item.get("searched_substance"),
                    item.get("substance"),
                    substance,
                ),
                "Strength": field_value(item, "strength", item.get("strength"), extract_strength(product)),
                "Dosage Form": field_value(
                    item,
                    "dosage_form",
                    item.get("dosage_form"),
                    extract_dosage_form(product),
                ),
                "Route": export_value(item.get("route"), extract_route(product)),
                "Pack Size": field_value(item, "pack_size", item.get("pack_size"), extract_pack_size(product)),
                "ATC Code": field_value(item, "atc_code", item.get("atc_code")),
                "Therapeutic Category": field_value(
                    item,
                    "therapeutic_category",
                    short_therapeutic_category(
                        item.get("therapeutic_category"),
                        item.get("searched_substance") or item.get("substance") or substance,
                        item.get("atc_code"),
                    ),
                    therapeutic_category_from_atc(item.get("atc_code")),
                    default_category,
                ),
                "MA Holder Name": field_value(item, "company", company),
                "Manufacturer Name": field_value(item, "manufacturer_name", manufacturer_name_value(item)),
                "Manufacturer Country": field_value(
                    item,
                    "manufacturer_country",
                    manufacturer_country_value(item),
                ),
                "Registration Status": export_value(item.get("status")),
                "Registration Number": export_value(
                    item.get("registration_number"),
                    extract_registration_number(product),
                ),
                "Registration Date": export_value(item.get("registration_date")),
                "Expiry Date": export_value(item.get("expiry_date")),
                "Product Details": export_value(product_url, source_url),
                "PDF URL": export_value(pdf_url),
                "SMPC URL": field_value(item, "smpc_url", smpc_url),
                "PIL URL": field_value(item, "pil_url", pil_url),
                "PIL / Assessment Report": export_value(pil_or_assessment_document_url),
                "Assessment Report URL": field_value(
                    item,
                    "assessment_report_url",
                    assessment_document_url,
                ),
                "Manufacturer Website": export_value(item.get("manufacturer_website"), product_url),
                "Manufacturer Contact Us Phone Number": export_value(item.get("manufacturer_phone")),
                "Manufacturer Contact Us Email ID": export_value(item.get("manufacturer_email")),
                "Box Artwork": export_value(item.get("box_artwork_url")),
                "Foil Artwork": export_value(item.get("foil_artwork_url")),
                "Insert / PIL artwork": export_value(
                    item.get("insert_pil_artwork_url"),
                    pil_or_assessment_document_url,
                ),
                "SMPC": export_value(smpc_url),
                "Source": export_value(item.get("source")),
                "Source URL": export_value(source_url),
                "Document Type": export_value(infer_document_type(item)),
                "Data Confidence": export_value(item.get("data_confidence"), ai_metadata["data_confidence"]),
                "Enrichment Status": export_value(
                    item.get("enrichment_status"),
                    ai_metadata["enrichment_status"],
                ),
                "Missing Fields": export_value(item.get("missing_fields"), ai_metadata["missing_fields"]),
                "AI Next Action": export_value(item.get("ai_next_action"), ai_metadata["ai_next_action"]),
                "Last Checked": export_value(item.get("last_checked"), generated_at),
            }
        )
    return rows


def _coverage_rows(results: list[dict[str, Any]]) -> list[dict[str, object]]:
    fields = [
        ("Strength", "strength"),
        ("Dosage Form", "dosage_form"),
        ("Route", "route"),
        ("Pack Size", "pack_size"),
        ("ATC Code", "atc_code"),
        ("Therapeutic Category", "therapeutic_category"),
        ("SMPC URL", "smpc_url"),
        ("PIL URL", "pil_url"),
        ("Assessment Report", "assessment_report_url"),
        ("Expiry Date", "expiry_date"),
    ]
    by_source = {}
    default_category = default_therapeutic_category("", results)
    for item in results:
        item = english_row(item)
        source = item.get("source") or "Unknown"
        bucket = by_source.setdefault(
            source,
            {
                "Source": source,
                "Records": 0,
                **{label: 0 for label, _ in fields},
            },
        )
        bucket["Records"] += 1
        product = item.get("product", "")
        for label, field in fields:
            value = item.get(field)
            if not value and field == "strength":
                value = extract_strength(product)
            elif not value and field == "dosage_form":
                value = extract_dosage_form(product)
            elif not value and field == "route":
                value = extract_route(product)
            elif not value and field == "pack_size":
                value = extract_pack_size(product)
            elif not value and field == "therapeutic_category":
                value = first_value(
                    therapeutic_category_from_atc(item.get("atc_code")),
                    default_category,
                )
            if value:
                bucket[label] += 1

    rows = []
    for source, bucket in sorted(by_source.items()):
        records = bucket["Records"] or 1
        row = {"Source": source, "Records": bucket["Records"]}
        for label, _ in fields:
            row[label] = f"{bucket[label]}/{records}"
        rows.append(row)
    if not rows:
        rows.append({"Source": MISSING_VALUE, "Records": 0})
    return rows


def write_excel_export(substance: str, results: list[dict[str, Any]], deep: bool = False) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = "_deep" if deep else ""
    filename = f"{safe_filename(substance)}{suffix}_{timestamp}.xlsx"
    path = EXPORT_DIR / filename
    export_rows = build_export_rows(substance, results)
    df = pd.DataFrame(export_rows, columns=EXPORT_COLUMNS)
    provenance = pd.DataFrame(
        [
            {
                "Query": substance,
                "Generated At": datetime.now().isoformat(timespec="seconds"),
                "Record Count": len(results),
                "Sources": ", ".join(
                    sorted({item.get("source", "") for item in results if item.get("source")})
                )
                or MISSING_VALUE,
                "Status": "Records found" if results else "No records found",
            }
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Products")
        provenance.to_excel(writer, index=False, sheet_name="Provenance")
        if deep:
            pd.DataFrame(_coverage_rows(results)).to_excel(
                writer,
                index=False,
                sheet_name="Coverage",
            )
        worksheet = writer.sheets["Products"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        url_columns = {
            "Product Details",
            "PDF URL",
            "SMPC URL",
            "PIL URL",
            "PIL / Assessment Report",
            "Assessment Report URL",
            "Manufacturer Website",
            "Source URL",
        }
        name_columns = {
            "Product Name",
            "Brand Name",
            "Molecule",
            "Molecule (Active Ingredient(s))",
            "MA Holder Name",
            "Manufacturer Name",
        }
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap_alignment
        for column_index, column_cells in enumerate(worksheet.columns, start=1):
            header = str(column_cells[0].value or "")
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            if header in url_columns:
                width = min(max(max_length + 2, 28), 95)
            elif header in name_columns:
                width = min(max(max_length + 2, 18), 70)
            else:
                width = min(max(max_length + 2, 12), 42)
            column_letter = get_column_letter(column_index)
            worksheet.column_dimensions[column_letter].width = width
            for cell in column_cells:
                cell.alignment = wrap_alignment
        worksheet.row_dimensions[1].height = 34
        for row in range(2, min(worksheet.max_row, 250) + 1):
            worksheet.row_dimensions[row].height = 42
        for sheet_name, sheet in writer.sheets.items():
            if sheet_name == "Products":
                continue
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column_index, column_cells in enumerate(sheet.columns, start=1):
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_index)].width = min(
                    max(max_length + 2, 12),
                    60,
                )
    logger.info("Wrote %s export with %s records to %s", "deep" if deep else "standard", len(results), path)
    return Path(path)
